from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File, status, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime, timedelta
import io
import openpyxl
from pydantic import BaseModel

from models import (
    LoginRequest, TokenResponse, UserOut,
    Simpatisan, SimpatisanBase, Kader, KaderBase, Saksi, SaksiBase,
    WilayahTarget, WilayahTargetBase, QuickCount, QuickCountBase, uid
)
from auth import (
    hash_password, verify_password, create_access_token, get_current_user
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="SiPekaeS API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# -------- helper --------
def clean(doc):
    if doc:
        doc.pop('_id', None)
        doc.pop('hashed_password', None)
    return doc

def clean_list(items):
    return [clean(d) for d in items]

def area_filter(current: dict):
    """Return Mongo filter based on user role for koordinator/saksi."""
    role = current.get('role')
    if role in ('super_admin', 'admin_pusat', 'admin_input'):
        return {}
    # koordinator or saksi: fetch profile
    return None  # will be resolved async via helper below

async def get_area_filter(current: dict):
    role = current.get('role')
    if role in ('super_admin', 'admin_pusat', 'admin_input'):
        return {}
    user = await db.users.find_one({"username": current['sub']})
    if not user:
        return {}
    q = {}
    if role == 'koordinator':
        if user.get('kecamatan_kerja'): q['kecamatan'] = user['kecamatan_kerja']
        if user.get('desa_kerja'): q['desa'] = user['desa_kerja']
    elif role == 'saksi':
        if user.get('kecamatan_kerja'): q['kecamatan'] = user['kecamatan_kerja']
        if user.get('tps_kerja'): q['tps'] = user['tps_kerja']
    return q

# ================ ROOT ================
@api_router.get("/")
async def root():
    return {"message": "SiPekaeS API v2.0"}

# ================ AUTH ================
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(payload: LoginRequest):
    user = await db.users.find_one({"username": payload.username})
    if not user or not verify_password(payload.password, user['hashed_password']):
        raise HTTPException(status_code=401, detail="Username atau password salah")
    token = create_access_token({"sub": user['username'], "id": user['id'], "role": user['role']})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": UserOut(
            id=user['id'], username=user['username'], name=user['name'],
            role=user['role'], roleLabel=user.get('roleLabel'), avatar=user.get('avatar'),
            kecamatan_kerja=user.get('kecamatan_kerja', ''),
            desa_kerja=user.get('desa_kerja', ''),
            tps_kerja=user.get('tps_kerja', ''),
        )
    }

@api_router.get("/auth/me", response_model=UserOut)
async def me(current=Depends(get_current_user)):
    user = await db.users.find_one({"username": current['sub']})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return UserOut(
        id=user['id'], username=user['username'], name=user['name'],
        role=user['role'], roleLabel=user.get('roleLabel'), avatar=user.get('avatar'),
        kecamatan_kerja=user.get('kecamatan_kerja', ''),
        desa_kerja=user.get('desa_kerja', ''),
        tps_kerja=user.get('tps_kerja', ''),
    )

@api_router.get("/users", response_model=List[UserOut])
async def list_users(current=Depends(get_current_user)):
    users = await db.users.find().to_list(1000)
    return [UserOut(
        id=u['id'], username=u['username'], name=u['name'],
        role=u['role'], roleLabel=u.get('roleLabel'), avatar=u.get('avatar'),
        kecamatan_kerja=u.get('kecamatan_kerja', ''),
        desa_kerja=u.get('desa_kerja', ''),
        tps_kerja=u.get('tps_kerja', ''),
    ) for u in users]

class UserCreatePayload(LoginRequest):
    name: str
    role: str
    roleLabel: Optional[str] = ""
    avatar: Optional[str] = ""
    kecamatan_kerja: Optional[str] = ""
    desa_kerja: Optional[str] = ""
    tps_kerja: Optional[str] = ""

class UserUpdatePayload(LoginRequest):
    name: str
    role: str
    roleLabel: Optional[str] = ""
    avatar: Optional[str] = ""
    kecamatan_kerja: Optional[str] = ""
    desa_kerja: Optional[str] = ""
    tps_kerja: Optional[str] = ""
    password: Optional[str] = None  # optional (leave empty to not change)

@api_router.post("/users", response_model=UserOut)
async def create_user(payload: UserCreatePayload, current=Depends(get_current_user)):
    if current.get('role') not in ('super_admin', 'admin_pusat'):
        raise HTTPException(status_code=403, detail="Tidak berhak menambah pengguna")
    existing = await db.users.find_one({"username": payload.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username sudah dipakai")
    from models import uid as _uid
    doc = {
        'id': _uid(), 'username': payload.username, 'name': payload.name,
        'role': payload.role, 'roleLabel': payload.roleLabel or payload.role,
        'avatar': payload.avatar or '',
        'kecamatan_kerja': payload.kecamatan_kerja or '',
        'desa_kerja': payload.desa_kerja or '',
        'tps_kerja': payload.tps_kerja or '',
        'hashed_password': hash_password(payload.password),
        'created_at': datetime.utcnow(),
    }
    await db.users.insert_one(doc.copy())
    return UserOut(**{k: doc[k] for k in ['id','username','name','role','roleLabel','avatar','kecamatan_kerja','desa_kerja','tps_kerja']})

@api_router.put("/users/{user_id}", response_model=UserOut)
async def update_user(user_id: str, payload: UserUpdatePayload, current=Depends(get_current_user)):
    if current.get('role') not in ('super_admin', 'admin_pusat'):
        raise HTTPException(status_code=403, detail="Tidak berhak")
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    update = {
        'username': payload.username, 'name': payload.name, 'role': payload.role,
        'roleLabel': payload.roleLabel or payload.role, 'avatar': payload.avatar or '',
        'kecamatan_kerja': payload.kecamatan_kerja or '',
        'desa_kerja': payload.desa_kerja or '',
        'tps_kerja': payload.tps_kerja or '',
    }
    if payload.password:
        update['hashed_password'] = hash_password(payload.password)
    await db.users.update_one({"id": user_id}, {"$set": update})
    u = await db.users.find_one({"id": user_id})
    return UserOut(id=u['id'], username=u['username'], name=u['name'], role=u['role'],
                   roleLabel=u.get('roleLabel'), avatar=u.get('avatar'),
                   kecamatan_kerja=u.get('kecamatan_kerja',''), desa_kerja=u.get('desa_kerja',''), tps_kerja=u.get('tps_kerja',''))

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current=Depends(get_current_user)):
    if current.get('role') != 'super_admin':
        raise HTTPException(status_code=403, detail="Hanya Super Admin yang bisa hapus pengguna")
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    if user.get('username') == current.get('sub'):
        raise HTTPException(status_code=400, detail="Tidak bisa menghapus akun sendiri")
    await db.users.delete_one({"id": user_id})
    return {"ok": True}

class ChangePasswordPayload(BaseModel):
    old_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_password(payload: ChangePasswordPayload, current=Depends(get_current_user)):
    user = await db.users.find_one({"username": current['sub']})
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    if not verify_password(payload.old_password, user['hashed_password']):
        raise HTTPException(status_code=400, detail="Password lama salah")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password baru minimal 6 karakter")
    await db.users.update_one({"id": user['id']}, {"$set": {"hashed_password": hash_password(payload.new_password)}})
    return {"ok": True, "message": "Password berhasil diubah"}

@api_router.post("/admin/reset-demo-data")
async def reset_demo_data(current=Depends(get_current_user)):
    """Hapus SEMUA data (simpatisan, kader, saksi, wilayah_target) — TIDAK menghapus user.
    Hanya Super Admin yang bisa jalankan. Untuk siapkan aplikasi ke production tanpa data demo."""
    if current.get('role') != 'super_admin':
        raise HTTPException(status_code=403, detail="Hanya Super Admin")
    counts = {}
    for coll in ('simpatisan', 'kader', 'saksi', 'wilayah_target'):
        result = await db[coll].delete_many({})
        counts[coll] = result.deleted_count
    return {"ok": True, "deleted": counts, "message": "Semua data demo terhapus. User tetap tersimpan."}

# ================ CRUD Simpatisan / Kader / Saksi ================
def make_crud(prefix: str, collection: str, base_model, full_model):
    router = APIRouter()

    @router.get(f"/{prefix}")
    async def list_items(current=Depends(get_current_user)):
        q = await get_area_filter(current)
        items = await db[collection].find(q).sort("tanggal", -1).to_list(10000)
        return clean_list(items)

    @router.post(f"/{prefix}")
    async def create_item(data: base_model, current=Depends(get_current_user)):
        doc = full_model(**data.dict()).dict()
        await db[collection].insert_one(doc.copy())
        return clean(doc)

    @router.get(f"/{prefix}/{{item_id}}")
    async def get_item(item_id: str, current=Depends(get_current_user)):
        doc = await db[collection].find_one({"id": item_id})
        if not doc:
            raise HTTPException(status_code=404, detail="Not found")
        return clean(doc)

    @router.put(f"/{prefix}/{{item_id}}")
    async def update_item(item_id: str, data: base_model, current=Depends(get_current_user)):
        result = await db[collection].update_one({"id": item_id}, {"$set": data.dict()})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Not found")
        doc = await db[collection].find_one({"id": item_id})
        return clean(doc)

    @router.delete(f"/{prefix}/{{item_id}}")
    async def delete_item(item_id: str, current=Depends(get_current_user)):
        result = await db[collection].delete_one({"id": item_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Not found")
        return {"ok": True}

    return router

api_router.include_router(make_crud("simpatisan", "simpatisan", SimpatisanBase, Simpatisan))
api_router.include_router(make_crud("kader", "kader", KaderBase, Kader))
api_router.include_router(make_crud("saksi", "saksi", SaksiBase, Saksi))

# ================ AGGREGATED ORGANISASI ================
@api_router.get("/organisasi/{jenis}")
async def get_organisasi(jenis: str, current=Depends(get_current_user)):
    """Aggregate people from simpatisan+kader+saksi where they have specific organizational role flag."""
    flag_map = {
        'dpc': ('is_pengurus_dpc', 'jabatan_dpc'),
        'dpra': ('is_pengurus_dpra', 'jabatan_dpra'),
        'pelopor': ('is_pelopor', 'peran_pelopor'),
        'rki': ('is_rki', 'jabatan_rki'),
    }
    if jenis not in flag_map:
        raise HTTPException(status_code=404, detail="Jenis organisasi tidak dikenali")
    flag_field, jabatan_field = flag_map[jenis]
    base_area = await get_area_filter(current)
    q = {**base_area, flag_field: True}
    result = []
    for coll_name, source_label in [("simpatisan", "Simpatisan"), ("kader", "Kader"), ("saksi", "Saksi")]:
        async for d in db[coll_name].find(q):
            d.pop('_id', None)
            d['source_type'] = coll_name
            d['source_label'] = source_label
            d['jabatan_organisasi'] = d.get(jabatan_field, '')
            result.append(d)
    # sort by nama
    result.sort(key=lambda x: (x.get('nama') or '').lower())
    return result

# ================ WILAYAH TARGET ================
async def _compute_realisasi_map():
    """Compute unique-people count per kecamatan (Kader > Saksi > Simpatisan dedup)."""
    def key_of(d):
        nik = (d.get('nik') or '').strip()
        if nik: return f"NIK:{nik}"
        return f"NM:{(d.get('nama') or '').lower().strip()}|{(d.get('kecamatan') or '').lower().strip()}"
    kaders = await db.kader.find().to_list(50000)
    saksis = await db.saksi.find().to_list(50000)
    simpatisans = await db.simpatisan.find().to_list(50000)
    kader_keys = {key_of(k) for k in kaders}
    saksi_keys = {key_of(s) for s in saksis} - kader_keys
    simpatisan_keys = {key_of(s) for s in simpatisans} - kader_keys - saksi_keys
    counts = {}
    for k in kaders:
        kec = k.get('kecamatan')
        if kec: counts[kec] = counts.get(kec, 0) + 1
    for s in saksis:
        if key_of(s) not in saksi_keys: continue
        kec = s.get('kecamatan')
        if kec: counts[kec] = counts.get(kec, 0) + 1
    for sp in simpatisans:
        if key_of(sp) not in simpatisan_keys: continue
        kec = sp.get('kecamatan')
        if kec: counts[kec] = counts.get(kec, 0) + 1
    return counts

@api_router.get("/wilayah-target")
async def list_wilayah_target(current=Depends(get_current_user)):
    items = await db.wilayah_target.find().sort("kecamatan", 1).to_list(1000)
    realisasi_map = await _compute_realisasi_map()
    result = []
    for w in items:
        w = clean(w)
        w['realisasi'] = realisasi_map.get(w.get('kecamatan'), 0)
        result.append(w)
    return result

@api_router.post("/wilayah-target")
async def create_wilayah_target(data: WilayahTargetBase, current=Depends(get_current_user)):
    existing = await db.wilayah_target.find_one({"kecamatan": data.kecamatan})
    if existing:
        await db.wilayah_target.update_one(
            {"kecamatan": data.kecamatan},
            {"$set": {**data.dict(), "updated_at": datetime.utcnow()}}
        )
        doc = await db.wilayah_target.find_one({"kecamatan": data.kecamatan})
        return clean(doc)
    doc = WilayahTarget(**data.dict()).dict()
    await db.wilayah_target.insert_one(doc.copy())
    return clean(doc)

@api_router.put("/wilayah-target/{item_id}")
async def update_wilayah_target(item_id: str, data: WilayahTargetBase, current=Depends(get_current_user)):
    result = await db.wilayah_target.update_one(
        {"id": item_id},
        {"$set": {**data.dict(), "updated_at": datetime.utcnow()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    doc = await db.wilayah_target.find_one({"id": item_id})
    return clean(doc)

@api_router.delete("/wilayah-target/{item_id}")
async def delete_wilayah_target(item_id: str, current=Depends(get_current_user)):
    result = await db.wilayah_target.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}

# ================ QUICK COUNT ================
@api_router.get("/quick-count")
async def list_quick_count(current=Depends(get_current_user)):
    q = await get_area_filter(current)
    # For saksi role, further restrict to own tps
    if current.get('role') == 'saksi':
        user = await db.users.find_one({"username": current['sub']})
        if user and user.get('tps_kerja'):
            q['tps'] = user['tps_kerja']
    items = await db.quick_count.find(q).sort("submitted_at", -1).to_list(20000)
    return clean_list(items)

@api_router.post("/quick-count")
async def create_quick_count(data: QuickCountBase, current=Depends(get_current_user)):
    # Upsert per (kecamatan|tps) - satu TPS satu entry
    existing = await db.quick_count.find_one({"kecamatan": data.kecamatan, "tps": data.tps})
    if existing:
        update = {**data.dict(), "submitted_by": current.get('sub'), "submitted_at": datetime.utcnow()}
        await db.quick_count.update_one({"id": existing['id']}, {"$set": update})
        doc = await db.quick_count.find_one({"id": existing['id']})
        return clean(doc)
    doc = QuickCount(**data.dict(), submitted_by=current.get('sub','')).dict()
    await db.quick_count.insert_one(doc.copy())
    return clean(doc)

@api_router.put("/quick-count/{item_id}")
async def update_quick_count(item_id: str, data: QuickCountBase, current=Depends(get_current_user)):
    result = await db.quick_count.update_one(
        {"id": item_id},
        {"$set": {**data.dict(), "submitted_by": current.get('sub'), "submitted_at": datetime.utcnow()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    doc = await db.quick_count.find_one({"id": item_id})
    return clean(doc)

@api_router.delete("/quick-count/{item_id}")
async def delete_quick_count(item_id: str, current=Depends(get_current_user)):
    if current.get('role') not in ('super_admin', 'admin_pusat'):
        raise HTTPException(status_code=403, detail="Hanya admin yang bisa hapus")
    result = await db.quick_count.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}

@api_router.get("/quick-count/summary")
async def quick_count_summary(current=Depends(get_current_user)):
    """Aggregate quick count results — total suara per paslon + persentase."""
    items = await db.quick_count.find().to_list(50000)
    # Total TPS covered
    total_tps = len(items)
    # Count all saksi entries in system (target TPS)
    target_tps = await db.saksi.count_documents({})
    total = {"paslon_1": 0, "paslon_2": 0, "paslon_3": 0, "tidak_sah": 0, "dpt": 0}
    per_kec = {}
    for it in items:
        total["paslon_1"] += it.get('paslon_1', 0) or 0
        total["paslon_2"] += it.get('paslon_2', 0) or 0
        total["paslon_3"] += it.get('paslon_3', 0) or 0
        total["tidak_sah"] += it.get('suara_tidak_sah', 0) or 0
        total["dpt"] += it.get('dpt', 0) or 0
        kec = it.get('kecamatan', 'Lain')
        e = per_kec.setdefault(kec, {"kecamatan": kec, "paslon_1": 0, "paslon_2": 0, "paslon_3": 0, "tidak_sah": 0, "tps_terlapor": 0})
        e["paslon_1"] += it.get('paslon_1', 0) or 0
        e["paslon_2"] += it.get('paslon_2', 0) or 0
        e["paslon_3"] += it.get('paslon_3', 0) or 0
        e["tidak_sah"] += it.get('suara_tidak_sah', 0) or 0
        e["tps_terlapor"] += 1

    total_sah = total["paslon_1"] + total["paslon_2"] + total["paslon_3"]
    total_semua = total_sah + total["tidak_sah"]

    def pct(v, base):
        return round((v / base) * 100, 2) if base else 0

    return {
        "total_tps_terlapor": total_tps,
        "target_tps": max(target_tps, total_tps),
        "coverage_persen": pct(total_tps, max(target_tps, 1)),
        "total_suara_sah": total_sah,
        "total_suara_tidak_sah": total["tidak_sah"],
        "total_suara": total_semua,
        "total_dpt": total["dpt"],
        "partisipasi_persen": pct(total_semua, total["dpt"]) if total["dpt"] else 0,
        "paslon": [
            {"nama": "Paslon 1", "suara": total["paslon_1"], "persen": pct(total["paslon_1"], total_sah), "warna": "#EF4444"},
            {"nama": "Paslon 2 (Kami)", "suara": total["paslon_2"], "persen": pct(total["paslon_2"], total_sah), "warna": "#F97316"},
            {"nama": "Paslon 3", "suara": total["paslon_3"], "persen": pct(total["paslon_3"], total_sah), "warna": "#3B82F6"},
        ],
        "per_kecamatan": sorted(per_kec.values(), key=lambda x: x['kecamatan']),
    }

# ================ STATS ================
@api_router.get("/stats/summary")
async def stats_summary(current=Depends(get_current_user)):
    area = await get_area_filter(current)
    total_simpatisan = await db.simpatisan.count_documents(area)
    total_kader = await db.kader.count_documents(area)
    total_saksi = await db.saksi.count_documents(area)

    total_dpc = 0
    total_dpra = 0
    total_pelopor = 0
    total_rki = 0
    for coll_name in ("simpatisan", "kader", "saksi"):
        total_dpc += await db[coll_name].count_documents({**area, "is_pengurus_dpc": True})
        total_dpra += await db[coll_name].count_documents({**area, "is_pengurus_dpra": True})
        total_pelopor += await db[coll_name].count_documents({**area, "is_pelopor": True})
        total_rki += await db[coll_name].count_documents({**area, "is_rki": True})

    # RW tercover
    rw_set = set()
    for coll_name in ("simpatisan", "kader", "saksi"):
        async for d in db[coll_name].find({**area, "rw": {"$nin": ["", None]}}, {"kecamatan":1, "desa":1, "rw":1}):
            rw_set.add(f"{d.get('kecamatan')}|{d.get('desa')}|{d.get('rw')}")

    rw_tercover = len(rw_set)
    rw_total = 3000

    # Aggregate baseline/target/realisasi (realisasi auto from unique dedup count)
    total_baseline = 0
    total_target = 0
    realisasi_map = await _compute_realisasi_map()
    async for w in db.wilayah_target.find():
        total_baseline += w.get('baseline', 0) or 0
        total_target += w.get('target', 0) or 0
    total_realisasi = sum(realisasi_map.values())

    # growth (last 30 days)
    since = datetime.utcnow() - timedelta(days=30)
    growth_simpatisan = await db.simpatisan.count_documents({**area, "tanggal": {"$gte": since}})
    growth_kader = await db.kader.count_documents({**area, "tanggal": {"$gte": since}})
    growth_saksi = await db.saksi.count_documents({**area, "tanggal": {"$gte": since}})

    return {
        "simpatisan": {"value": total_simpatisan, "growth": growth_simpatisan},
        "kader": {"value": total_kader, "growth": growth_kader},
        "saksi": {"value": total_saksi, "growth": growth_saksi},
        "pengurus_dpc": total_dpc,
        "pengurus_dpra": total_dpra,
        "pelopor": total_pelopor,
        "rki": total_rki,
        "rw": {"value": rw_tercover, "total": rw_total, "tercover": round((rw_tercover / rw_total) * 100, 1) if rw_total else 0},
        "kecamatan": 47,
        "desa": 381,
        "baseline": total_baseline,
        "target": total_target if total_target else 1,
        "realisasi": total_realisasi,
    }

@api_router.get("/stats/per-kecamatan")
async def stats_per_kecamatan(current=Depends(get_current_user)):
    result = {}
    for coll_name, key in [("simpatisan", "simpatisan"), ("kader", "kader"), ("saksi", "saksi")]:
        async for row in db[coll_name].aggregate([{"$group": {"_id": "$kecamatan", "count": {"$sum": 1}}}]):
            kec = row['_id']
            if kec:
                result.setdefault(kec, {}).update({key: row['count']})
    # merge with wilayah_target
    async for w in db.wilayah_target.find():
        kec = w.get('kecamatan')
        if kec:
            result.setdefault(kec, {}).update({
                "baseline": w.get('baseline', 0),
                "target": w.get('target', 0),
                "realisasi": w.get('realisasi', 0),
            })
    return result

@api_router.get("/stats/kecamatan-detail")
async def stats_kecamatan_detail(current=Depends(get_current_user)):
    """Per-kecamatan detail with dedup (Kader > Saksi > Simpatisan).
    Also counts DPC/DPRA/Pelopor/RKI members per kecamatan."""
    def key_of(d):
        nik = (d.get('nik') or '').strip()
        if nik:
            return f"NIK:{nik}"
        return f"NM:{(d.get('nama') or '').lower().strip()}|{(d.get('kecamatan') or '').lower().strip()}"

    # Load all people
    kaders = await db.kader.find().to_list(50000)
    saksis = await db.saksi.find().to_list(50000)
    simpatisans = await db.simpatisan.find().to_list(50000)

    kader_keys = {key_of(k) for k in kaders}
    saksi_keys = {key_of(s) for s in saksis} - kader_keys  # saksi minus kader
    simpatisan_keys = {key_of(s) for s in simpatisans} - kader_keys - saksi_keys

    # Aggregate per kecamatan
    by_kec = {}
    for k in kaders:
        kec = k.get('kecamatan')
        if not kec: continue
        entry = by_kec.setdefault(kec, {'kecamatan': kec, 'kader': 0, 'saksi': 0, 'simpatisan': 0, 'dpc': 0, 'dpra': 0, 'pelopor': 0, 'rki': 0, 'total_unik': 0})
        entry['kader'] += 1
        entry['total_unik'] += 1
        if k.get('is_pengurus_dpc'): entry['dpc'] += 1
        if k.get('is_pengurus_dpra'): entry['dpra'] += 1
        if k.get('is_pelopor'): entry['pelopor'] += 1
        if k.get('is_rki'): entry['rki'] += 1

    for s in saksis:
        if key_of(s) not in saksi_keys: continue
        kec = s.get('kecamatan')
        if not kec: continue
        entry = by_kec.setdefault(kec, {'kecamatan': kec, 'kader': 0, 'saksi': 0, 'simpatisan': 0, 'dpc': 0, 'dpra': 0, 'pelopor': 0, 'rki': 0, 'total_unik': 0})
        entry['saksi'] += 1
        entry['total_unik'] += 1
        if s.get('is_pengurus_dpc'): entry['dpc'] += 1
        if s.get('is_pengurus_dpra'): entry['dpra'] += 1
        if s.get('is_pelopor'): entry['pelopor'] += 1
        if s.get('is_rki'): entry['rki'] += 1

    for sp in simpatisans:
        if key_of(sp) not in simpatisan_keys: continue
        kec = sp.get('kecamatan')
        if not kec: continue
        entry = by_kec.setdefault(kec, {'kecamatan': kec, 'kader': 0, 'saksi': 0, 'simpatisan': 0, 'dpc': 0, 'dpra': 0, 'pelopor': 0, 'rki': 0, 'total_unik': 0})
        entry['simpatisan'] += 1
        entry['total_unik'] += 1
        if sp.get('is_pengurus_dpc'): entry['dpc'] += 1
        if sp.get('is_pengurus_dpra'): entry['dpra'] += 1
        if sp.get('is_pelopor'): entry['pelopor'] += 1
        if sp.get('is_rki'): entry['rki'] += 1

    # merge with wilayah_target (baseline/target from stored, realisasi = total_unik)
    async for w in db.wilayah_target.find():
        kec = w.get('kecamatan')
        if not kec: continue
        entry = by_kec.setdefault(kec, {'kecamatan': kec, 'kader': 0, 'saksi': 0, 'simpatisan': 0, 'dpc': 0, 'dpra': 0, 'pelopor': 0, 'rki': 0, 'total_unik': 0})
        entry['baseline'] = w.get('baseline', 0)
        entry['target'] = w.get('target', 0)
    # realisasi = total_unik (auto)
    for e in by_kec.values():
        e['realisasi'] = e['total_unik']

    return sorted(by_kec.values(), key=lambda x: x['kecamatan'])

@api_router.get("/stats/desa-detail")
async def stats_desa_detail(current=Depends(get_current_user)):
    """Aggregate per desa/kelurahan with dedup and RW list."""
    def key_of(d):
        nik = (d.get('nik') or '').strip()
        if nik: return f"NIK:{nik}"
        return f"NM:{(d.get('nama') or '').lower().strip()}|{(d.get('kecamatan') or '').lower().strip()}"

    kaders = await db.kader.find().to_list(50000)
    saksis = await db.saksi.find().to_list(50000)
    simpatisans = await db.simpatisan.find().to_list(50000)

    kader_keys = {key_of(k) for k in kaders}
    saksi_keys = {key_of(s) for s in saksis} - kader_keys
    simpatisan_keys = {key_of(s) for s in simpatisans} - kader_keys - saksi_keys

    by_desa = {}
    def desa_key(d):
        return f"{d.get('kecamatan','')}|{d.get('desa','')}"

    def ensure(d):
        dk = desa_key(d)
        if not d.get('desa'): return None
        e = by_desa.setdefault(dk, {
            'kecamatan': d.get('kecamatan',''), 'desa': d.get('desa',''),
            'kader': 0, 'saksi': 0, 'simpatisan': 0, 'total': 0,
            'rws': set(),
        })
        if d.get('rw'): e['rws'].add(d['rw'])
        return e

    for k in kaders:
        e = ensure(k)
        if e: e['kader'] += 1; e['total'] += 1
    for s in saksis:
        if key_of(s) not in saksi_keys: continue
        e = ensure(s)
        if e: e['saksi'] += 1; e['total'] += 1
    for sp in simpatisans:
        if key_of(sp) not in simpatisan_keys: continue
        e = ensure(sp)
        if e: e['simpatisan'] += 1; e['total'] += 1

    out = []
    for e in by_desa.values():
        out.append({**e, 'rws': sorted(e['rws']), 'rw_count': len(e['rws'])})
    return sorted(out, key=lambda x: (x['kecamatan'], x['desa']))

@api_router.get("/stats/rw-detail")
async def stats_rw_detail(current=Depends(get_current_user)):
    def key_of(d):
        nik = (d.get('nik') or '').strip()
        if nik: return f"NIK:{nik}"
        return f"NM:{(d.get('nama') or '').lower().strip()}|{(d.get('kecamatan') or '').lower().strip()}"

    kaders = await db.kader.find().to_list(50000)
    saksis = await db.saksi.find().to_list(50000)
    simpatisans = await db.simpatisan.find().to_list(50000)

    kader_keys = {key_of(k) for k in kaders}
    saksi_keys = {key_of(s) for s in saksis} - kader_keys
    simpatisan_keys = {key_of(s) for s in simpatisans} - kader_keys - saksi_keys

    by_rw = {}
    def rw_key(d):
        return f"{d.get('kecamatan','')}|{d.get('desa','')}|{d.get('rw','')}"
    def ensure(d):
        if not d.get('rw'): return None
        rk = rw_key(d)
        return by_rw.setdefault(rk, {
            'kecamatan': d.get('kecamatan',''), 'desa': d.get('desa',''), 'rw': d.get('rw',''),
            'kader': 0, 'saksi': 0, 'simpatisan': 0, 'total': 0,
        })
    for k in kaders:
        e = ensure(k)
        if e: e['kader'] += 1; e['total'] += 1
    for s in saksis:
        if key_of(s) not in saksi_keys: continue
        e = ensure(s)
        if e: e['saksi'] += 1; e['total'] += 1
    for sp in simpatisans:
        if key_of(sp) not in simpatisan_keys: continue
        e = ensure(sp)
        if e: e['simpatisan'] += 1; e['total'] += 1
    return sorted(by_rw.values(), key=lambda x: (x['kecamatan'], x['desa'], x['rw']))

@api_router.get("/stats/daily-growth")
async def daily_growth(days: int = 30, current=Depends(get_current_user)):
    """Return daily counts of new simpatisan/kader/saksi for last N days + last update timestamps."""
    from datetime import date
    days = max(1, min(90, days))
    now = datetime.utcnow()
    since = now - timedelta(days=days - 1)
    # Bucket by date
    buckets = {}
    for i in range(days):
        d = (since + timedelta(days=i)).date().isoformat()
        buckets[d] = {"date": d, "simpatisan": 0, "kader": 0, "saksi": 0}

    last_update = {"simpatisan": None, "kader": None, "saksi": None}

    for coll_name, key in [("simpatisan", "simpatisan"), ("kader", "kader"), ("saksi", "saksi")]:
        # Latest doc for last_update
        latest = await db[coll_name].find_one(sort=[("tanggal", -1)])
        if latest and latest.get('tanggal'):
            last_update[key] = latest['tanggal'].isoformat() if hasattr(latest['tanggal'], 'isoformat') else latest['tanggal']
        # Bucket counts
        cursor = db[coll_name].find({"tanggal": {"$gte": since}}, {"tanggal": 1})
        async for d in cursor:
            t = d.get('tanggal')
            if not t: continue
            key_d = t.date().isoformat() if hasattr(t, 'date') else str(t)[:10]
            if key_d in buckets:
                buckets[key_d][key] += 1

    series = list(buckets.values())
    # Add cumulative totals for each row (running total)
    cum = {"simpatisan": 0, "kader": 0, "saksi": 0}
    # Get starting totals (records before `since`)
    for coll_name, key in [("simpatisan", "simpatisan"), ("kader", "kader"), ("saksi", "saksi")]:
        cum[key] = await db[coll_name].count_documents({"tanggal": {"$lt": since}})
    for row in series:
        cum["simpatisan"] += row["simpatisan"]
        cum["kader"] += row["kader"]
        cum["saksi"] += row["saksi"]
        row["total_simpatisan"] = cum["simpatisan"]
        row["total_kader"] = cum["kader"]
        row["total_saksi"] = cum["saksi"]
        row["total_semua"] = cum["simpatisan"] + cum["kader"] + cum["saksi"]

    # Latest last_update across all collections
    latest_ts = max([v for v in last_update.values() if v] or [now.isoformat()])

    return {
        "series": series,
        "last_update": {
            "simpatisan": last_update["simpatisan"],
            "kader": last_update["kader"],
            "saksi": last_update["saksi"],
            "latest": latest_ts,
        },
        "totals": {
            "simpatisan_baru_30h": sum(r["simpatisan"] for r in series),
            "kader_baru_30h": sum(r["kader"] for r in series),
            "saksi_baru_30h": sum(r["saksi"] for r in series),
        },
    }

# ================ EXCEL IMPORT / EXPORT ================
@api_router.get("/simpatisan/template/excel")
async def download_template(current=Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Simpatisan"
    headers = ["nama", "nik", "hp", "kecamatan", "desa", "rw", "rt", "alamat"]
    ws.append(headers)
    ws.append(["Contoh Nama", "3202010000000001", "081234567890", "Cikembar", "Mekarjaya", "RW 04", "RT 02", "Jl. Contoh No. 1"])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_simpatisan.xlsx"}
    )

@api_router.post("/simpatisan/import/excel")
async def import_excel(file: UploadFile = File(...), current=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file: {str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File kosong")

    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    for r in ('nama', 'kecamatan'):
        if r not in headers:
            raise HTTPException(status_code=400, detail=f"Kolom '{r}' wajib ada")

    inserted = 0
    errors = []
    docs = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(row): continue
        try:
            data = {}
            for h, v in zip(headers, row):
                if h and v is not None: data[h] = str(v).strip()
            if not data.get('nama') or not data.get('kecamatan'):
                errors.append(f"Baris {i}: nama/kecamatan kosong"); continue
            doc = Simpatisan(
                nama=data.get('nama',''), nik=data.get('nik',''), hp=data.get('hp',''),
                kecamatan=data.get('kecamatan',''), desa=data.get('desa',''),
                rw=data.get('rw',''), rt=data.get('rt',''), alamat=data.get('alamat',''),
                status='aktif',
            ).dict()
            docs.append(doc); inserted += 1
        except Exception as e:
            errors.append(f"Baris {i}: {str(e)}")

    if docs: await db.simpatisan.insert_many(docs)
    return {"inserted": inserted, "errors": errors[:20], "total_errors": len(errors)}

# ================ Excel Kader & Saksi ================
def _make_template(fields, sample):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(fields)
    ws.append(sample)
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

@api_router.get("/kader/template/excel")
async def kader_template(current=Depends(get_current_user)):
    buf = _make_template(
        ["nama", "jabatan", "hp", "kecamatan", "desa", "rw", "alamat"],
        ["Contoh Kader", "Kader Aktif", "081234567890", "Cikembar", "Mekarjaya", "RW 04", "Jl. Contoh No. 1"]
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_kader.xlsx"})

@api_router.post("/kader/import/excel")
async def kader_import(file: UploadFile = File(...), current=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus .xlsx")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True); ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal baca: {e}")
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2: raise HTTPException(status_code=400, detail="File kosong")
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    for r in ('nama', 'kecamatan', 'jabatan'):
        if r not in headers: raise HTTPException(status_code=400, detail=f"Kolom '{r}' wajib ada")
    inserted, errors, docs = 0, [], []
    for i, row in enumerate(rows[1:], start=2):
        if not any(row): continue
        try:
            data = {h: str(v).strip() for h, v in zip(headers, row) if h and v is not None}
            if not data.get('nama') or not data.get('kecamatan') or not data.get('jabatan'):
                errors.append(f"Baris {i}: nama/kecamatan/jabatan wajib"); continue
            docs.append(Kader(
                nama=data.get('nama',''), jabatan=data.get('jabatan',''), hp=data.get('hp',''),
                kecamatan=data.get('kecamatan',''), desa=data.get('desa',''), rw=data.get('rw',''),
                alamat=data.get('alamat','')
            ).dict())
            inserted += 1
        except Exception as e:
            errors.append(f"Baris {i}: {e}")
    if docs: await db.kader.insert_many(docs)
    return {"inserted": inserted, "errors": errors[:20], "total_errors": len(errors)}

@api_router.get("/saksi/template/excel")
async def saksi_template(current=Depends(get_current_user)):
    buf = _make_template(
        ["nama", "tps", "hp", "kecamatan", "desa", "rw", "alamat", "status"],
        ["Contoh Saksi", "TPS 01", "081234567890", "Cikembar", "Mekarjaya", "RW 04", "Jl. Contoh No. 1", "pending"]
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_saksi.xlsx"})

@api_router.post("/saksi/import/excel")
async def saksi_import(file: UploadFile = File(...), current=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus .xlsx")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True); ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal baca: {e}")
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2: raise HTTPException(status_code=400, detail="File kosong")
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    for r in ('nama', 'kecamatan', 'tps'):
        if r not in headers: raise HTTPException(status_code=400, detail=f"Kolom '{r}' wajib ada")
    inserted, errors, docs = 0, [], []
    for i, row in enumerate(rows[1:], start=2):
        if not any(row): continue
        try:
            data = {h: str(v).strip() for h, v in zip(headers, row) if h and v is not None}
            if not data.get('nama') or not data.get('kecamatan') or not data.get('tps'):
                errors.append(f"Baris {i}: nama/kecamatan/tps wajib"); continue
            docs.append(Saksi(
                nama=data.get('nama',''), tps=data.get('tps',''), hp=data.get('hp',''),
                kecamatan=data.get('kecamatan',''), desa=data.get('desa',''), rw=data.get('rw',''),
                alamat=data.get('alamat',''), status=data.get('status','pending')
            ).dict())
            inserted += 1
        except Exception as e:
            errors.append(f"Baris {i}: {e}")
    if docs: await db.saksi.insert_many(docs)
    return {"inserted": inserted, "errors": errors[:20], "total_errors": len(errors)}

# ================ Include router ================
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()
