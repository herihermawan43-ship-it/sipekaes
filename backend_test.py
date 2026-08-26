#!/usr/bin/env python3
"""
Comprehensive Backend API Test Suite for SiPekaeS
Tests all endpoints: Auth, CRUD for 7 entities, Stats, Excel import/export
"""
import requests
import json
import io
import openpyxl
from typing import Dict, Any, Optional

# Read backend URL from frontend/.env
with open('/app/frontend/.env', 'r') as f:
    for line in f:
        if line.startswith('REACT_APP_BACKEND_URL='):
            BASE_URL = line.split('=')[1].strip()
            break

API_URL = f"{BASE_URL}/api"

# Test credentials
TEST_CREDS = {
    "username": "superadmin",
    "password": "admin123"
}

# Global token storage
AUTH_TOKEN = None

# Test results tracking
test_results = {
    "passed": [],
    "failed": [],
    "warnings": []
}

def log_test(name: str, passed: bool, message: str = ""):
    """Log test result"""
    status = "✅ PASS" if passed else "❌ FAIL"
    print(f"{status}: {name}")
    if message:
        print(f"   → {message}")
    
    if passed:
        test_results["passed"].append(name)
    else:
        test_results["failed"].append({"test": name, "message": message})

def log_warning(name: str, message: str):
    """Log warning (non-critical issue)"""
    print(f"⚠️  WARNING: {name}")
    print(f"   → {message}")
    test_results["warnings"].append({"test": name, "message": message})

# ============ AUTH TESTS ============
def test_login_valid():
    """Test login with valid credentials"""
    global AUTH_TOKEN
    try:
        response = requests.post(f"{API_URL}/auth/login", json=TEST_CREDS, timeout=10)
        
        if response.status_code != 200:
            log_test("Login with valid credentials", False, f"Expected 200, got {response.status_code}: {response.text}")
            return False
        
        data = response.json()
        
        # Check response structure
        if "access_token" not in data:
            log_test("Login with valid credentials", False, "Missing access_token in response")
            return False
        
        if "user" not in data:
            log_test("Login with valid credentials", False, "Missing user object in response")
            return False
        
        user = data["user"]
        if user.get("role") != "super_admin":
            log_test("Login with valid credentials", False, f"Expected role=super_admin, got {user.get('role')}")
            return False
        
        if user.get("username") != "superadmin":
            log_test("Login with valid credentials", False, f"Expected username=superadmin, got {user.get('username')}")
            return False
        
        AUTH_TOKEN = data["access_token"]
        log_test("Login with valid credentials", True, f"Token received, user role: {user.get('role')}")
        return True
        
    except Exception as e:
        log_test("Login with valid credentials", False, f"Exception: {str(e)}")
        return False

def test_login_invalid():
    """Test login with invalid credentials"""
    try:
        response = requests.post(f"{API_URL}/auth/login", 
                                json={"username": "superadmin", "password": "wrongpassword"}, 
                                timeout=10)
        
        if response.status_code == 401:
            log_test("Login with invalid credentials", True, "Correctly returned 401")
            return True
        else:
            log_test("Login with invalid credentials", False, f"Expected 401, got {response.status_code}")
            return False
            
    except Exception as e:
        log_test("Login with invalid credentials", False, f"Exception: {str(e)}")
        return False

def test_auth_me():
    """Test /auth/me endpoint with valid token"""
    if not AUTH_TOKEN:
        log_test("GET /auth/me with token", False, "No auth token available")
        return False
    
    try:
        headers = {"Authorization": f"Bearer {AUTH_TOKEN}"}
        response = requests.get(f"{API_URL}/auth/me", headers=headers, timeout=10)
        
        if response.status_code != 200:
            log_test("GET /auth/me with token", False, f"Expected 200, got {response.status_code}: {response.text}")
            return False
        
        user = response.json()
        if user.get("username") != "superadmin":
            log_test("GET /auth/me with token", False, f"Expected username=superadmin, got {user.get('username')}")
            return False
        
        log_test("GET /auth/me with token", True, f"User: {user.get('name')}, Role: {user.get('role')}")
        return True
        
    except Exception as e:
        log_test("GET /auth/me with token", False, f"Exception: {str(e)}")
        return False

def test_auth_me_no_token():
    """Test /auth/me endpoint without token (should return 401)"""
    try:
        response = requests.get(f"{API_URL}/auth/me", timeout=10)
        
        if response.status_code == 401:
            log_test("GET /auth/me without token", True, "Correctly returned 401")
            return True
        else:
            log_test("GET /auth/me without token", False, f"Expected 401, got {response.status_code}")
            return False
            
    except Exception as e:
        log_test("GET /auth/me without token", False, f"Exception: {str(e)}")
        return False

def test_list_users():
    """Test GET /users endpoint"""
    if not AUTH_TOKEN:
        log_test("GET /users", False, "No auth token available")
        return False
    
    try:
        headers = {"Authorization": f"Bearer {AUTH_TOKEN}"}
        response = requests.get(f"{API_URL}/users", headers=headers, timeout=10)
        
        if response.status_code != 200:
            log_test("GET /users", False, f"Expected 200, got {response.status_code}: {response.text}")
            return False
        
        users = response.json()
        if not isinstance(users, list):
            log_test("GET /users", False, "Response is not a list")
            return False
        
        if len(users) < 5:
            log_test("GET /users", False, f"Expected at least 5 users, got {len(users)}")
            return False
        
        log_test("GET /users", True, f"Retrieved {len(users)} users")
        return True
        
    except Exception as e:
        log_test("GET /users", False, f"Exception: {str(e)}")
        return False

# ============ CRUD TESTS ============
def test_crud_entity(prefix: str, sample_data: Dict[str, Any]):
    """Test full CRUD operations for an entity"""
    if not AUTH_TOKEN:
        print(f"⚠️  Skipping CRUD tests for {prefix}: No auth token")
        return False
    
    headers = {"Authorization": f"Bearer {AUTH_TOKEN}"}
    created_id = None
    all_passed = True
    
    # Test 1: GET list (should work even if empty)
    try:
        response = requests.get(f"{API_URL}/{prefix}", headers=headers, timeout=10)
        if response.status_code == 200:
            items = response.json()
            log_test(f"GET /api/{prefix} (list)", True, f"Retrieved {len(items)} items")
        else:
            log_test(f"GET /api/{prefix} (list)", False, f"Status {response.status_code}: {response.text}")
            all_passed = False
    except Exception as e:
        log_test(f"GET /api/{prefix} (list)", False, f"Exception: {str(e)}")
        all_passed = False
    
    # Test 2: POST create
    try:
        response = requests.post(f"{API_URL}/{prefix}", json=sample_data, headers=headers, timeout=10)
        if response.status_code == 200:
            created = response.json()
            if "id" in created:
                created_id = created["id"]
                log_test(f"POST /api/{prefix} (create)", True, f"Created with id: {created_id[:8]}...")
            else:
                log_test(f"POST /api/{prefix} (create)", False, "Response missing 'id' field")
                all_passed = False
        else:
            log_test(f"POST /api/{prefix} (create)", False, f"Status {response.status_code}: {response.text}")
            all_passed = False
    except Exception as e:
        log_test(f"POST /api/{prefix} (create)", False, f"Exception: {str(e)}")
        all_passed = False
    
    # Test 3: GET single (if we have an ID)
    if created_id:
        try:
            response = requests.get(f"{API_URL}/{prefix}/{created_id}", headers=headers, timeout=10)
            if response.status_code == 200:
                item = response.json()
                log_test(f"GET /api/{prefix}/{{id}} (single)", True, f"Retrieved item {created_id[:8]}...")
            else:
                log_test(f"GET /api/{prefix}/{{id}} (single)", False, f"Status {response.status_code}: {response.text}")
                all_passed = False
        except Exception as e:
            log_test(f"GET /api/{prefix}/{{id}} (single)", False, f"Exception: {str(e)}")
            all_passed = False
    
    # Test 4: PUT update (if we have an ID)
    if created_id:
        try:
            update_data = sample_data.copy()
            # Update a field based on entity type
            if "nama" in update_data:
                update_data["nama"] = update_data["nama"] + " (Updated)"
            
            response = requests.put(f"{API_URL}/{prefix}/{created_id}", json=update_data, headers=headers, timeout=10)
            if response.status_code == 200:
                updated = response.json()
                log_test(f"PUT /api/{prefix}/{{id}} (update)", True, f"Updated item {created_id[:8]}...")
            else:
                log_test(f"PUT /api/{prefix}/{{id}} (update)", False, f"Status {response.status_code}: {response.text}")
                all_passed = False
        except Exception as e:
            log_test(f"PUT /api/{prefix}/{{id}} (update)", False, f"Exception: {str(e)}")
            all_passed = False
    
    # Test 5: DELETE (if we have an ID)
    if created_id:
        try:
            response = requests.delete(f"{API_URL}/{prefix}/{created_id}", headers=headers, timeout=10)
            if response.status_code == 200:
                result = response.json()
                if result.get("ok") == True:
                    log_test(f"DELETE /api/{prefix}/{{id}}", True, f"Deleted item {created_id[:8]}...")
                else:
                    log_test(f"DELETE /api/{prefix}/{{id}}", False, f"Response missing 'ok: true': {result}")
                    all_passed = False
            else:
                log_test(f"DELETE /api/{prefix}/{{id}}", False, f"Status {response.status_code}: {response.text}")
                all_passed = False
        except Exception as e:
            log_test(f"DELETE /api/{prefix}/{{id}}", False, f"Exception: {str(e)}")
            all_passed = False
    
    # Test 6: Verify 401 without token
    try:
        response = requests.get(f"{API_URL}/{prefix}", timeout=10)
        if response.status_code == 401:
            log_test(f"GET /api/{prefix} without token (401 check)", True, "Correctly returned 401")
        else:
            log_test(f"GET /api/{prefix} without token (401 check)", False, f"Expected 401, got {response.status_code}")
            all_passed = False
    except Exception as e:
        log_test(f"GET /api/{prefix} without token (401 check)", False, f"Exception: {str(e)}")
        all_passed = False
    
    return all_passed

# ============ STATS TESTS ============
def test_stats_summary():
    """Test GET /stats/summary endpoint"""
    if not AUTH_TOKEN:
        log_test("GET /stats/summary", False, "No auth token available")
        return False
    
    try:
        headers = {"Authorization": f"Bearer {AUTH_TOKEN}"}
        response = requests.get(f"{API_URL}/stats/summary", headers=headers, timeout=10)
        
        if response.status_code != 200:
            log_test("GET /stats/summary", False, f"Expected 200, got {response.status_code}: {response.text}")
            return False
        
        stats = response.json()
        
        # Check required fields
        required_fields = ["simpatisan", "kader", "saksi", "rw", "kecamatan", "desa"]
        missing = [f for f in required_fields if f not in stats]
        if missing:
            log_test("GET /stats/summary", False, f"Missing fields: {missing}")
            return False
        
        # Check RW structure
        if "rw" in stats:
            rw = stats["rw"]
            if not isinstance(rw, dict) or "tercover" not in rw:
                log_test("GET /stats/summary", False, f"RW field missing 'tercover' percentage")
                return False
        
        log_test("GET /stats/summary", True, 
                f"Simpatisan: {stats.get('simpatisan', {}).get('value', 0)}, "
                f"Kader: {stats.get('kader', {}).get('value', 0)}, "
                f"RW tercover: {stats.get('rw', {}).get('tercover', 0)}%")
        return True
        
    except Exception as e:
        log_test("GET /stats/summary", False, f"Exception: {str(e)}")
        return False

# ============ EXCEL TESTS ============
def test_excel_template_download():
    """Test GET /simpatisan/template/excel"""
    if not AUTH_TOKEN:
        log_test("GET /simpatisan/template/excel", False, "No auth token available")
        return False
    
    try:
        headers = {"Authorization": f"Bearer {AUTH_TOKEN}"}
        response = requests.get(f"{API_URL}/simpatisan/template/excel", headers=headers, timeout=10)
        
        if response.status_code != 200:
            log_test("GET /simpatisan/template/excel", False, f"Expected 200, got {response.status_code}: {response.text}")
            return False
        
        # Check content type
        content_type = response.headers.get('content-type', '')
        if 'spreadsheet' not in content_type and 'excel' not in content_type:
            log_warning("GET /simpatisan/template/excel", f"Content-Type is '{content_type}', expected Excel MIME type")
        
        # Check if it's a valid Excel file (starts with PK for zip format)
        content = response.content
        if not content.startswith(b'PK'):
            log_test("GET /simpatisan/template/excel", False, "Response is not a valid Excel file (doesn't start with PK)")
            return False
        
        # Try to load it with openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers_row = [cell.value for cell in ws[1]]
            log_test("GET /simpatisan/template/excel", True, f"Valid Excel file with headers: {headers_row}")
            return True
        except Exception as e:
            log_test("GET /simpatisan/template/excel", False, f"Failed to parse Excel: {str(e)}")
            return False
        
    except Exception as e:
        log_test("GET /simpatisan/template/excel", False, f"Exception: {str(e)}")
        return False

def test_excel_import():
    """Test POST /simpatisan/import/excel"""
    if not AUTH_TOKEN:
        log_test("POST /simpatisan/import/excel", False, "No auth token available")
        return False
    
    try:
        # Create a test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Import"
        
        # Headers
        headers = ["nama", "nik", "hp", "kecamatan", "desa", "rw", "rt", "alamat"]
        ws.append(headers)
        
        # Sample data rows
        ws.append(["Budi Santoso", "3202011234567890", "081234567890", "Cikembar", "Mekarjaya", "RW 04", "RT 02", "Jl. Merdeka No. 10"])
        ws.append(["Siti Nurhaliza", "3202012345678901", "081298765432", "Cicurug", "Sukamaju", "RW 03", "RT 01", "Jl. Raya Sukabumi No. 25"])
        ws.append(["Ahmad Hidayat", "3202013456789012", "081387654321", "Cibadak", "Cibadak Kota", "RW 05", "RT 03", "Jl. Pahlawan No. 5"])
        
        # Save to BytesIO
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        # Upload
        headers = {"Authorization": f"Bearer {AUTH_TOKEN}"}
        files = {"file": ("test_simpatisan.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = requests.post(f"{API_URL}/simpatisan/import/excel", files=files, headers=headers, timeout=10)
        
        if response.status_code != 200:
            log_test("POST /simpatisan/import/excel", False, f"Expected 200, got {response.status_code}: {response.text}")
            return False
        
        result = response.json()
        
        # Check response structure
        if "inserted" not in result:
            log_test("POST /simpatisan/import/excel", False, "Response missing 'inserted' field")
            return False
        
        if "errors" not in result:
            log_test("POST /simpatisan/import/excel", False, "Response missing 'errors' field")
            return False
        
        inserted = result["inserted"]
        errors = result["errors"]
        
        if inserted < 3:
            log_test("POST /simpatisan/import/excel", False, f"Expected 3 insertions, got {inserted}. Errors: {errors}")
            return False
        
        log_test("POST /simpatisan/import/excel", True, f"Inserted {inserted} rows, {len(errors)} errors")
        return True
        
    except Exception as e:
        log_test("POST /simpatisan/import/excel", False, f"Exception: {str(e)}")
        return False

# ============ MAIN TEST RUNNER ============
def main():
    print("=" * 80)
    print("SiPekaeS Backend API Test Suite")
    print(f"Testing: {API_URL}")
    print("=" * 80)
    print()
    
    # Phase 1: Authentication Tests
    print("📋 PHASE 1: AUTHENTICATION TESTS")
    print("-" * 80)
    test_login_valid()
    test_login_invalid()
    test_auth_me()
    test_auth_me_no_token()
    test_list_users()
    print()
    
    # Phase 2: CRUD Tests for all entities
    print("📋 PHASE 2: CRUD TESTS FOR ALL ENTITIES")
    print("-" * 80)
    
    entities = [
        ("simpatisan", {
            "nama": "Rina Wijaya",
            "nik": "3202014567890123",
            "hp": "081276543210",
            "kecamatan": "Parungkuda",
            "desa": "Bojong Jengkol",
            "rw": "RW 02",
            "rt": "RT 04",
            "alamat": "Jl. Raya Parungkuda No. 15",
            "status": "aktif"
        }),
        ("kader", {
            "nama": "Dedi Mulyadi",
            "jabatan": "Ketua Kecamatan",
            "kecamatan": "Cikembar",
            "desa": "Mekarjaya",
            "rw": "RW 01",
            "hp": "081365432109",
            "alamat": "Jl. Kader No. 8"
        }),
        ("saksi", {
            "nama": "Ani Rahmawati",
            "tps": "TPS 001",
            "kecamatan": "Cicurug",
            "desa": "Sukamaju",
            "rw": "RW 03",
            "hp": "081454321098",
            "status": "terverifikasi"
        }),
        ("pengurus-dpc", {
            "nama": "Hendra Gunawan",
            "jabatan": "Ketua DPC",
            "hp": "081543210987",
            "alamat": "Jl. DPC Sukabumi No. 1",
            "foto": ""
        }),
        ("pengurus-dpra", {
            "nama": "Lina Marlina",
            "jabatan": "Sekretaris DPRA",
            "kecamatan": "Cibadak",
            "desa": "Cibadak Kota",
            "hp": "081632109876",
            "kategori": "kader"
        }),
        ("pelopor", {
            "nama": "Yudi Setiawan",
            "kecamatan": "Parungkuda",
            "desa": "Bojong Jengkol",
            "hp": "081721098765",
            "peran": "Koordinator"
        }),
        ("rki", {
            "nama": "Eko Prasetyo",
            "jabatan": "Ketua RKI",
            "kecamatan": "Cikembar",
            "desa": "Mekarjaya",
            "hp": "081810987654"
        })
    ]
    
    for prefix, sample_data in entities:
        print(f"\n--- Testing {prefix.upper()} ---")
        test_crud_entity(prefix, sample_data)
    
    print()
    
    # Phase 3: Stats Tests
    print("📋 PHASE 3: STATS TESTS")
    print("-" * 80)
    test_stats_summary()
    print()
    
    # Phase 4: Excel Tests
    print("📋 PHASE 4: EXCEL IMPORT/EXPORT TESTS")
    print("-" * 80)
    test_excel_template_download()
    test_excel_import()
    print()
    
    # Summary
    print("=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    print(f"✅ Passed: {len(test_results['passed'])}")
    print(f"❌ Failed: {len(test_results['failed'])}")
    print(f"⚠️  Warnings: {len(test_results['warnings'])}")
    
    if test_results['failed']:
        print("\n❌ FAILED TESTS:")
        for fail in test_results['failed']:
            print(f"  - {fail['test']}")
            print(f"    {fail['message']}")
    
    if test_results['warnings']:
        print("\n⚠️  WARNINGS:")
        for warn in test_results['warnings']:
            print(f"  - {warn['test']}")
            print(f"    {warn['message']}")
    
    print("\n" + "=" * 80)
    
    # Exit code
    if test_results['failed']:
        exit(1)
    else:
        exit(0)

if __name__ == "__main__":
    main()
